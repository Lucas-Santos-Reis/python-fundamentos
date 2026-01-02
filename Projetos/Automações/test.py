from openpyxl import load_workbook
arquivo = load_workbook('Alunos.xlsx') #carrega o arquivo excel
print(arquivo.sheetnames) #mostra as abas do excel
aba_ativa = arquivo.active #pega a aba ativa
aba_alunos = arquivo['Planilha1'] #pega a aba especifica
# selecionar células
valor_a1 = aba_alunos["A1"].value
valor_b1 = aba_alunos.cell(row=1, column=2).value

aba_alunos.cell(row=1, column=2).value = "Prova 1"

arquivo.save("Alunos2.xlsx") #salva a modificação em outro arquivo

# ultima linha
print(aba_alunos.max_row)
print(len(aba_alunos["A"]))
# percorrer toda a nossa base de dados
# para cada item
    #   ver se o bairro já existe em uma aba, se não existir, criar aquela aba
    #   copiar os valores daquela linha e colocar na aba do bairro correspondente
